import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import pandas as pd

def generate_excel(df, shift_mapping, prices, holidays, worker_name="N/D", company_name="N/D"):
    """
    Genera un Excel con:
    1. Pestaña "RESUMEN EJECUTIVO".
    2. Pestaña por cada Mes "ENERO", "FEBRERO", etc.
    """
    output = BytesIO()
    wb = openpyxl.Workbook()
    
    # --- ESTILOS COMUNES ---
    COLOR_HEADER_BG = "4F81BD"  # Azul solicitado
    COLOR_HEADER_FONT = "FFFFFF"
    
    # Estilo Rojo (Deuda)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006", bold=True)

    # Estilo Verde (Festivo)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100", bold=True)
    
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    border_medium = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    
    font_bold = Font(bold=True)
    font_title = Font(bold=True, size=14)
    font_subtitle = Font(bold=True, size=12)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # ---------------------------------------------------------
    # HOJA 1: RESUMEN EJECUTIVO (MEJORADO)
    # ---------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "RESUMEN EJECUTIVO"
    
    # 1. Cabecera del Informe
    ws_summary.merge_cells('B2:E2')
    ws_summary['B2'] = "INFORME TÉCNICO V3 - AUDITORÍA & RECLAMACIÓN"
    ws_summary['B2'].font = font_title
    ws_summary['B2'].alignment = align_center
    ws_summary['B2'].border = border_medium
    
    # Datos Extraídos (Parser Auditor)
    cat_prof = prices.get('categoria', 'N/D')
    
    ws_summary['B4'] = "TRABAJADOR:"
    ws_summary['C4'] = worker_name
    ws_summary['B5'] = "CATEGORÍA PROF:"
    ws_summary['C5'] = cat_prof
    ws_summary['B6'] = "EMPRESA:"
    ws_summary['C6'] = company_name
    ws_summary['B7'] = "FECHA INFORME:"
    ws_summary['C7'] = pd.Timestamp.now().strftime("%d/%m/%Y")
    
    for row in range(4, 8):
        ws_summary[f'B{row}'].font = font_bold
        ws_summary[f'C{row}'].alignment = Alignment(horizontal='left')
        
    # 2. CÁLCULO DEL PRECIO HORA ORDINARIA (FÓRMULA VISUAL)
    # Según imagen: (Base + Antigüedad + Plus) x 15 / 1776
    
    # Extraer valores
    base = prices.get('base_salary', 0.0)
    antiguedad = prices.get('seniority', 0.0)
    plus = prices.get('plus_agreement', 0.0)
    
    # Calcular totales internos para consistencia (aunque el usuario pase el precio ya calculado, 
    # aquí mostramos el desglose "teórico" de su fórmula anual)
    total_mensual = base + antiguedad + plus
    total_anual = total_mensual * 15
    divisor_horas = 1776
    precio_hora_formula = total_anual / divisor_horas if divisor_horas else 0.0
    
    # Cabecera Sección
    ws_summary.merge_cells('B9:E9')
    ws_summary['B9'] = "CÁLCULO DEL PRECIO HORA ORDINARIA"
    ws_summary['B9'].font = font_subtitle
    ws_summary['B9'].alignment = align_center
    ws_summary['B9'].border = border_thin
    
    # Fila Fórmula Texto
    # "(1253.26 + 100.26 + 0.00) x 15"
    formula_text = f"({base:.2f} + {antiguedad:.2f} + {plus:.2f}) x 15"
    ws_summary.merge_cells('B11:E11')
    ws_summary['B11'] = formula_text
    ws_summary['B11'].alignment = align_center
    ws_summary['B11'].font = Font(size=12, bold=True)
    
    # Fila Línea Divisoria
    ws_summary.merge_cells('B12:E12')
    ws_summary['B12'] = "--------------------------------------------------"
    ws_summary['B12'].alignment = align_center
    
    # Fila Divisor
    ws_summary.merge_cells('B13:E13')
    ws_summary['B13'] = str(divisor_horas)
    ws_summary['B13'].alignment = align_center
    ws_summary['B13'].font = Font(size=12, bold=True)
    
    # Resultado
    ws_summary.merge_cells('B15:C15')
    ws_summary['B15'] = "PRECIO HORA:"
    ws_summary['B15'].font = font_bold
    ws_summary['B15'].alignment = align_right
    
    ws_summary.merge_cells('D15:E15')
    ws_summary['D15'] = precio_hora_formula
    ws_summary['D15'].number_format = '#,##0.0000 €'
    ws_summary['D15'].font = Font(color="0000FF", bold=True, size=12)
    ws_summary['D15'].alignment = align_center

    # Actualizar el precio usado para cálculos posteriores si queremos que coincida exactamente
    # OJO: El main.py pasa 'price_normal'. Si el usuario quiere ESTA fórmula, 
    # deberíamos usar este precio. Pero para no romper lógica externa, solo visualizamos aquí.
    # El importe en detalle mensual usa 'precio_hora' que viene de main.
    
    # IMPORTANTE: Definimos precio_hora aquí para que el loop mensual lo use
    precio_hora = precio_hora_formula
    
    # 3. RESUMEN DE CANTIDADES (CABECERA)
    ws_summary.merge_cells('B18:E18')
    ws_summary['B18'] = "RESUMEN DE CANTIDADES A RECLAMAR"
    ws_summary['B18'].font = font_subtitle
    ws_summary['B18'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type='solid')
    ws_summary['B18'].border = border_thin
    
    # Totales
    total_deuda_descansos = 0.0 # Se calculará sumando el DF, o pasamos el total en 'prices'?
    # En la versión actual codigo, prices tiene 'val_extra_pay'. 
    # El total de descansos se calcula en Main. Deberíamos pasarlo o recalcularlo aquí.
    # Recalculamos rápido del DF para ser precisos
    total_horas_deuda = df['Deuda_Descanso_Horas'].sum()
    # Usamos el PRECIO DE FÓRMULA (1776)
    importe_descansos = total_horas_deuda * precio_hora 
    reclamacion_extra = prices.get('val_extra_pay', 0.0)
    total_final = importe_descansos + reclamacion_extra
    
    # Filas Resumen
    rows_summary = [
        ("Total Horas de Descanso No Disfrutadas:", f"{total_horas_deuda:.2f} h"),
        ("Importe Reclamación Descansos:", importe_descansos),
        ("Reclamación 3ª Paga Extra:", reclamacion_extra)
    ]
    
    curr = 20
    for label, val in rows_summary:
        ws_summary.merge_cells(f'B{curr}:C{curr}')
        ws_summary[f'B{curr}'] = label
        
        ws_summary.merge_cells(f'D{curr}:E{curr}')
        if isinstance(val, (int, float)):
             ws_summary[f'D{curr}'] = val
             ws_summary[f'D{curr}'].number_format = '#,##0.00 €'
        else:
             ws_summary[f'D{curr}'] = val
             ws_summary[f'D{curr}'].alignment = align_right
        
        ws_summary[f'D{curr}'].font = font_bold
        curr += 1
        
    # TOTAL FINAL (AMARILLO)
    curr += 1
    ws_summary.merge_cells(f'B{curr}:C{curr}')
    ws_summary[f'B{curr}'] = "TOTAL FINAL A RECLAMAR"
    ws_summary[f'B{curr}'].font = Font(bold=True, color="FF0000")
    ws_summary[f'B{curr}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
    ws_summary[f'B{curr}'].border = border_thin
    
    ws_summary.merge_cells(f'D{curr}:E{curr}')
    ws_summary[f'D{curr}'] = total_final
    ws_summary[f'D{curr}'].number_format = '#,##0.00 €'
    ws_summary[f'D{curr}'].font = Font(bold=True, color="FF0000", size=12)
    ws_summary[f'D{curr}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type='solid')
    ws_summary[f'D{curr}'].border = border_thin
    
    # ---------------------------------------------------------
    # HOJA 2: DETALLE MENSUAL (BLOQUES VISUALES)
    # ---------------------------------------------------------
    
    # Preparar datos
    df['Mes_Num'] = df['Fecha'].apply(lambda x: x.month)
    meses_ordenados = sorted(df['Mes_Num'].unique())
    
    month_names = {
        1:"ENERO", 2:"FEBRERO", 3:"MARZO", 4:"ABRIL", 5:"MAYO", 6:"JUNIO", 
        7:"JULIO", 8:"AGOSTO", 9:"SEPTIEMBRE", 10:"OCTUBRE", 11:"NOVIEMBRE", 12:"DICIEMBRE"
    }
    
    header_fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type='solid')
    header_font_style = Font(bold=True, color=COLOR_HEADER_FONT)
    holidays_set = set(holidays)

    # Crear única hoja
    ws_detail = wb.create_sheet("DETALLE MENSUAL")
    current_row = 1

    # Definir anchos
    widths = [12, 12, 12, 10, 10, 12, 15, 20, 15]
    for i, w in enumerate(widths, 1):
        ws_detail.column_dimensions[get_column_letter(i)].width = w

    for m_num in meses_ordenados:
        m_name = month_names.get(m_num, f"MES_{m_num}")
        
        # 1. SEPARADOR VISUAL (Bloque Azul)
        ws_detail.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        cell_title = ws_detail.cell(row=current_row, column=1, value=f"MES: {m_name}")
        cell_title.font = Font(bold=True, size=14, color="FFFFFF")
        cell_title.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type='solid')
        cell_title.alignment = align_center
        current_row += 1
        
        # 2. CABECERAS
        headers = ["FECHA", "ENTRADA", "SALIDA", "CÓDIGO", "HORAS", "DEUDA (H)", "TIEMPO DESCANSO", "ESTADO", "IMPORTE (€)"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=current_row, column=col_num, value=header)
            cell.font = header_font_style
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_thin
            
        current_row += 1
        
        # 3. DATOS
        df_month = df[df['Mes_Num'] == m_num].sort_values('Fecha')
        
        for _, row in df_month.iterrows():
            code = row.get('Codigo', '')
            info = shift_mapping.get(code, {})
            start = info.get('start', '-')
            end = info.get('end', '-')
            
            es_vacation = info.get('is_vacation')
            if es_vacation or code == 'V':
                start = "-"; end = "-"
                
            debt_h = row.get('Deuda_Descanso_Horas', 0.0)
            debt_minutes = int(round(debt_h * 60))
            if debt_h > 0:
                rest_str = f"{debt_minutes} min" 
            else:
                rest_str = "-"
            
            # Estado Logic
            date_obj = row['Fecha']
            is_holiday_manual = date_obj in holidays_set
            tipo = row.get('Tipo_Jornada', 'Ordinario')
            
            estado_texto = "Ordinario"
            if is_holiday_manual: estado_texto = "Festivo"
            elif tipo == "Festivo": estado_texto = "Festivo"
            elif es_vacation or code == 'V': estado_texto = "Vacaciones"
            elif info.get('type') == 'Absentismo': estado_texto = info.get('description', 'Absentismo')
            elif info.get('type') == 'Unknown': estado_texto = "Revisar"
            
            importe = debt_h * precio_hora
            
            # Escribir fila
            vals = [
                date_obj.strftime("%d/%m/%Y"), start, end, code, 
                row.get('Horas_Totales', 0), debt_h, rest_str, estado_texto, importe
            ]
            
            for c_idx, val in enumerate(vals, 1):
                cell = ws_detail.cell(row=current_row, column=c_idx, value=val)
                cell.alignment = align_center
                cell.border = border_thin
                
                # Format Importe
                if c_idx == 9: 
                    cell.number_format = '#,##0.00 €'
                    cell.alignment = align_right
                
                # Colores Condicionales
                
                # Definir amarillo
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                # 6: Deuda (H), 7: Tiempo Descanso -> ROJO si hay deuda
                if c_idx in [6, 7] and debt_h > 0:
                    cell.fill = red_fill
                    cell.font = red_font
                    
                # ESTADOS (Verde Festivo / Amarillo Vacaciones)
                if estado_texto == "Festivo":
                     # Toda la fila verde en texto o solo celda estado?
                     # Usuario dijo: "igual que festivo esta coloreado de verde... vacaciones amarillo"
                     # Si 'Festivo' ya se coloreaba:
                     if c_idx == 8: # Columna Estado
                         cell.fill = green_fill
                         cell.font = green_font
                     if c_idx == 1: # Fecha
                         cell.fill = green_fill
                         
                elif estado_texto == "Vacaciones":
                     # Aplicar Amarillo a toda la fila o celdas clave
                     # "vacaciones lo quiero coloreado de amarillo" -> Aplicaremos a Fecha, Codigo y Estado
                     if c_idx in [1, 4, 8]:
                         cell.fill = yellow_fill
                         cell.font = Font(color="000000", bold=True) # Texto negro para contraste

            current_row += 1
            
        # 4. ESPACIO EN BLANCO (Separador de Bloques)
        current_row += 2 # Dejar una fila vacía extra

    wb.save(output)
    output.seek(0)
    return output
